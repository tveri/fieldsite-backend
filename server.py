from glob import glob
import hashlib
from pprint import pprint
from sys import hash_info
from wsgiref.util import request_uri
from flask import Flask, request, send_from_directory, jsonify, render_template, redirect, make_response, send_file
import re, os, shutil, datetime, time, json, sqlite3, threading
import openpyxl as ox
import xml.dom.minidom as xml
from flask_cors import CORS, cross_origin
from openpyxl.utils.cell import get_column_letter
from bs4 import BeautifulSoup


MONTHS = [
    'Январь',
    'Февраль',
    'Март',
    'Апрель',
    'Май',
    'Июнь',
    'Июль',
    'Август',
    'Сентябрь',
    'Октябрь',
    'Ноябрь',
    'Декабрь',
]

app = Flask(__name__, static_url_path="")
app.config["DEBUG"] = True
app.config["JSON_AS_ASCII"] = False
app.config["HOST"] = "0.0.0.0"
application = app
CORS(app, support_credentials=True)

usersDb = sqlite3.connect('./users.db', check_same_thread=False)

db = sqlite3.connect('./allfields.db', check_same_thread=False)
cur = db.cursor()

fieldColumns = {col[0]: col[1] for col in cur.execute('PRAGMA table_info(ooo_kyrsk_agroaktiv__gorshechenskiy__field62_05)').fetchall()}
fieldColumnsNames = {fieldColumns[k]: k for k in fieldColumns}

fieldGlobalColumns = {col[0]: col[1] for col in cur.execute('PRAGMA table_info(ooo_kyrsk_agroaktiv)').fetchall()}
fieldGlobalColumnsNames = {fieldGlobalColumns[k]: k for k in fieldGlobalColumns}

lock = threading.Lock()

redYellowGreenFade = {0: 'ff0000', 1: 'ff0100', 2: 'ff0200', 3: 'ff0300', 4: 'ff0400', 5: 'ff0500', 6: 'ff0600', 7: 'ff0700', 8: 'ff0800', 9: 'ff0900', 10: 'ff0a00', 11: 'ff0b00', 12: 'ff0c00', 13: 'ff0d00', 14: 'ff0e00', 15: 'ff0f00', 16: 'ff1000', 17: 'ff1100', 18: 'ff1200', 19: 'ff1300', 20: 'ff1400', 21: 'ff1500', 22: 'ff1600', 23: 'ff1700', 24: 'ff1800', 25: 'ff1900', 26: 'ff1a00', 27: 'ff1b00', 28: 'ff1c00', 29: 'ff1d00', 30: 'ff1e00', 31: 'ff1f00', 32: 'ff2000', 33: 'ff2100', 34: 'ff2200', 35: 'ff2300', 36: 'ff2400', 37: 'ff2500', 38: 'ff2600', 39: 'ff2700', 40: 'ff2800', 41: 'ff2900', 42: 'ff2a00', 43: 'ff2b00', 44: 'ff2c00', 45: 'ff2d00', 46: 'ff2e00', 47: 'ff2f00', 48: 'ff3000', 49: 'ff3100', 50: 'ff3200', 51: 'ff3300', 52: 'ff3400', 53: 'ff3500', 54: 'ff3600', 55: 'ff3700', 56: 'ff3800', 57: 'ff3900', 58: 'ff3a00', 59: 'ff3b00', 60: 'ff3c00', 61: 'ff3d00', 62: 'ff3e00', 63: 'ff3f00', 64: 'ff4000', 65: 'ff4100', 66: 'ff4200', 67: 'ff4300', 68: 'ff4400', 69: 'ff4500', 70: 'ff4600', 71: 'ff4700', 72: 'ff4800', 73: 'ff4900', 74: 'ff4a00', 75: 'ff4b00', 76: 'ff4c00', 77: 'ff4d00', 78: 'ff4e00', 79: 'ff4f00', 80: 'ff5000', 81: 'ff5100', 82: 'ff5200', 83: 'ff5300', 84: 'ff5400', 85: 'ff5500', 86: 'ff5600', 87: 'ff5700', 88: 'ff5800', 89: 'ff5900', 90: 'ff5a00', 91: 'ff5b00', 92: 'ff5c00', 93: 'ff5d00', 94: 'ff5e00', 95: 'ff5f00', 96: 'ff6000', 97: 'ff6100', 98: 'ff6200', 99: 'ff6300', 100: 'ff6400', 101: 'ff6500', 102: 'ff6600', 103: 'ff6700', 104: 'ff6800', 105: 'ff6900', 106: 'ff6a00', 107: 'ff6b00', 108: 'ff6c00', 109: 'ff6d00', 110: 'ff6e00', 111: 'ff6f00', 112: 'ff7000', 113: 'ff7100', 114: 'ff7200', 115: 'ff7300', 116: 'ff7400', 117: 'ff7500', 118: 'ff7600', 119: 'ff7700', 120: 'ff7800', 121: 'ff7900', 122: 'ff7a00', 123: 'ff7b00', 124: 'ff7c00', 125: 'ff7d00', 126: 'ff7e00', 127: 'ff7f00', 128: 'ff8000', 129: 'ff8100', 130: 'ff8200', 131: 'ff8300', 132: 'ff8400', 133: 'ff8500', 134: 'ff8600', 135: 'ff8700', 136: 'ff8800', 137: 'ff8900', 138: 'ff8a00', 139: 'ff8b00', 140: 'ff8c00', 141: 'ff8d00', 142: 'ff8e00', 143: 'ff8f00', 144: 'ff9000', 145: 'ff9100', 146: 'ff9200', 147: 'ff9300', 148: 'ff9400', 149: 'ff9500', 150: 'ff9600', 151: 'ff9700', 152: 'ff9800', 153: 'ff9900', 154: 'ff9a00', 155: 'ff9b00', 156: 'ff9c00', 157: 'ff9d00', 158: 'ff9e00', 159: 'ff9f00', 160: 'ffa000', 161: 'ffa100', 162: 'ffa200', 163: 'ffa300', 164: 'ffa400', 165: 'ffa500', 166: 'ffa600', 167: 'ffa700', 168: 'ffa800', 169: 'ffa900', 170: 'ffaa00', 171: 'ffab00', 172: 'ffac00', 173: 'ffad00', 174: 'ffae00', 175: 'ffaf00', 176: 'ffb000', 177: 'ffb100', 178: 'ffb200', 179: 'ffb300', 180: 'ffb400', 181: 'ffb500', 182: 'ffb600', 183: 'ffb700', 184: 'ffb800', 185: 'ffb900', 186: 'ffba00', 187: 'ffbb00', 188: 'ffbc00', 189: 'ffbd00', 190: 'ffbe00', 191: 'ffbf00', 192: 'ffc000', 193: 'ffc100', 194: 'ffc200', 195: 'ffc300', 196: 'ffc400', 197: 'ffc500', 198: 'ffc600', 199: 'ffc700', 200: 'ffc800', 201: 'ffc900', 202: 'ffca00', 203: 'ffcb00', 204: 'ffcc00', 205: 'ffcd00', 206: 'ffce00', 207: 'ffcf00', 208: 'ffd000', 209: 'ffd100', 210: 'ffd200', 211: 'ffd300', 212: 'ffd400', 213: 'ffd500', 214: 'ffd600', 215: 'ffd700', 216: 'ffd800', 217: 'ffd900', 218: 'ffda00', 219: 'ffdb00', 220: 'ffdc00', 221: 'ffdd00', 222: 'ffde00', 223: 'ffdf00', 224: 'ffe000', 225: 'ffe100', 226: 'ffe200', 227: 'ffe300', 228: 'ffe400', 229: 'ffe500', 230: 'ffe600', 231: 'ffe700', 232: 'ffe800', 233: 'ffe900', 234: 'ffea00', 235: 'ffeb00', 236: 'ffec00', 237: 'ffed00', 238: 'ffee00', 239: 'ffef00', 240: 'fff000', 241: 'fff100', 242: 'fff200', 243: 'fff300', 244: 'fff400', 245: 'fff500', 246: 'fff600', 247: 'fff700', 248: 'fff800', 249: 'fff900', 250: 'fffa00', 251: 'fffb00', 252: 'fffc00', 253: 'fffd00', 254: 'fffe00', 255: 'ffff00', 256: 'ffff00', 257: 'feff00', 258: 'fdff00', 259: 'fcff00', 260: 'fbff00', 261: 'faff00', 262: 'f9ff00', 263: 'f8ff00', 264: 'f7ff00', 265: 'f6ff00', 266: 'f5ff00', 267: 'f4ff00', 268: 'f3ff00', 269: 'f2ff00', 270: 'f1ff00', 271: 'f0ff00', 272: 'efff00', 273: 'eeff00', 274: 'edff00', 275: 'ecff00', 276: 'ebff00', 277: 'eaff00', 278: 'e9ff00', 279: 'e8ff00', 280: 'e7ff00', 281: 'e6ff00', 282: 'e5ff00', 283: 'e4ff00', 284: 'e3ff00', 285: 'e2ff00', 286: 'e1ff00', 287: 'e0ff00', 288: 'dfff00', 289: 'deff00', 290: 'ddff00', 291: 'dcff00', 292: 'dbff00', 293: 'daff00', 294: 'd9ff00', 295: 'd8ff00', 296: 'd7ff00', 297: 'd6ff00', 298: 'd5ff00', 299: 'd4ff00', 300: 'd3ff00', 301: 'd2ff00', 302: 'd1ff00', 303: 'd0ff00', 304: 'cfff00', 305: 'ceff00', 306: 'cdff00', 307: 'ccff00', 308: 'cbff00', 309: 'caff00', 310: 'c9ff00', 311: 'c8ff00', 312: 'c7ff00', 313: 'c6ff00', 314: 'c5ff00', 315: 'c4ff00', 316: 'c3ff00', 317: 'c2ff00', 318: 'c1ff00', 319: 'c0ff00', 320: 'bfff00', 321: 'beff00', 322: 'bdff00', 323: 'bcff00', 324: 'bbff00', 325: 'baff00', 326: 'b9ff00', 327: 'b8ff00', 328: 'b7ff00', 329: 'b6ff00', 330: 'b5ff00', 331: 'b4ff00', 332: 'b3ff00', 333: 'b2ff00', 334: 'b1ff00', 335: 'b0ff00', 336: 'afff00', 337: 'aeff00', 338: 'adff00', 339: 'acff00', 340: 'abff00', 341: 'aaff00', 342: 'a9ff00', 343: 'a8ff00', 344: 'a7ff00', 345: 'a6ff00', 346: 'a5ff00', 347: 'a4ff00', 348: 'a3ff00', 349: 'a2ff00', 350: 'a1ff00', 351: 'a0ff00', 352: '9fff00', 353: '9eff00', 354: '9dff00', 355: '9cff00', 356: '9bff00', 357: '9aff00', 358: '99ff00', 359: '98ff00', 360: '97ff00', 361: '96ff00', 362: '95ff00', 363: '94ff00', 364: '93ff00', 365: '92ff00', 366: '91ff00', 367: '90ff00', 368: '8fff00', 369: '8eff00', 370: '8dff00', 371: '8cff00', 372: '8bff00', 373: '8aff00', 374: '89ff00', 375: '88ff00', 376: '87ff00', 377: '86ff00', 378: '85ff00', 379: '84ff00', 380: '83ff00', 381: '82ff00', 382: '81ff00', 383: '80ff00', 384: '7fff00', 385: '7eff00', 386: '7dff00', 387: '7cff00', 388: '7bff00', 389: '7aff00', 390: '79ff00', 391: '78ff00', 392: '77ff00', 393: '76ff00', 394: '75ff00', 395: '74ff00', 396: '73ff00', 397: '72ff00', 398: '71ff00', 399: '70ff00', 400: '6fff00', 401: '6eff00', 402: '6dff00', 403: '6cff00', 404: '6bff00', 405: '6aff00', 406: '69ff00', 407: '68ff00', 408: '67ff00', 409: '66ff00', 410: '65ff00', 411: '64ff00', 412: '63ff00', 413: '62ff00', 414: '61ff00', 415: '60ff00', 416: '5fff00', 417: '5eff00', 418: '5dff00', 419: '5cff00', 420: '5bff00', 421: '5aff00', 422: '59ff00', 423: '58ff00', 424: '57ff00', 425: '56ff00', 426: '55ff00', 427: '54ff00', 428: '53ff00', 429: '52ff00', 430: '51ff00', 431: '50ff00', 432: '4fff00', 433: '4eff00', 434: '4dff00', 435: '4cff00', 436: '4bff00', 437: '4aff00', 438: '49ff00', 439: '48ff00', 440: '47ff00', 441: '46ff00', 442: '45ff00', 443: '44ff00', 444: '43ff00', 445: '42ff00', 446: '41ff00', 447: '40ff00', 448: '3fff00', 449: '3eff00', 450: '3dff00', 451: '3cff00', 452: '3bff00', 453: '3aff00', 454: '39ff00', 455: '38ff00', 456: '37ff00', 457: '36ff00', 458: '35ff00', 459: '34ff00', 460: '33ff00', 461: '32ff00', 462: '31ff00', 463: '30ff00', 464: '2fff00', 465: '2eff00', 466: '2dff00', 467: '2cff00', 468: '2bff00', 469: '2aff00', 470: '29ff00', 471: '28ff00', 472: '27ff00', 473: '26ff00', 474: '25ff00', 475: '24ff00', 476: '23ff00', 477: '22ff00', 478: '21ff00', 479: '20ff00', 480: '1fff00', 481: '1eff00', 482: '1dff00', 483: '1cff00', 484: '1bff00', 485: '1aff00', 486: '19ff00', 487: '18ff00', 488: '17ff00', 489: '16ff00', 490: '15ff00', 491: '14ff00', 492: '13ff00', 493: '12ff00', 494: '11ff00', 495: '10ff00', 496: '0fff00', 497: '0eff00', 498: '0dff00', 499: '0cff00', 500: '0bff00', 501: '0aff00', 502: '09ff00', 503: '08ff00', 504: '07ff00', 505: '06ff00', 506: '05ff00', 507: '04ff00', 508: '03ff00', 509: '02ff00', 510: '01ff00', 511: '00ff00'}

writableColumnsMatch = {
    1: 0,
    2: 1,
    3: 2,
    4: 3,
    5: 4,
    10: 5,
    20: 6,
    21: 7,
}

def getConnection():
    return sqlite3.connect('./fields.db', check_same_thread=False)


def tryDecorator(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            # print('calc error: ', func, e)
            return 0
    return wrapper


@tryDecorator
def Cfunc(data, row, globalData):
    if data[row][1] < globalData[fieldGlobalColumnsNames['sowing_date']]:
        return 0
    return (data[row][1] - globalData[fieldGlobalColumnsNames['sowing_date']] + 86400) // 86400


@tryDecorator
def Gfunc(data, row):
    if data[row][3] > 0:
        return 0.0061 * ((25 + data[row][3]) ** 2) * (1 - 0.01 * data[row][4]) * 0.64 * (1 + 0.19 * data[row][5])
    return 0


@tryDecorator
def Hfunc(data, row):
    return 0.0000015592 * data[row][0] ** 3 - 0.0007462857 * data[row][0] ** 2 + 0.0896413572 * data[row][0] + 0.9787526972


@tryDecorator
def Ifunc(data, row):
    if data[row][2] > 0:
        return data[row][3] + (data[row-1][8] if row else 0)
    return 0


@tryDecorator
def Jfunc(data, row):
    if data[row][8] == 0:
        return 0
    return -0.000000000177496 * data[row][8] ** 3 + 0.000000336028555 * data[row][8] ** 2 + 0.0001264253108 * data[row][8] + 0.734104749417373


@tryDecorator
def Lfunc(data, row):
    if data[row][10] == 0:
        return 0.75 * data[row][6]
    return data[row][10] * data[row][6]


@tryDecorator
def Mfunc(data, row):
    if data[row][10] == 0:
        return 0.25
    return -0.000000000025495 * data[row][8] ** 3 - 0.000000045527389 * data[row][8] ** 2 + 0.000367679195804 * data[row][8] + 0.250076486014065


@tryDecorator
def Nfunc(data, row, globalData):
    return globalData[fieldGlobalColumnsNames['porisity']] * globalData[fieldGlobalColumnsNames['ppv']] * data[row][12] * 1000


@tryDecorator
def Ofunc(data, row, globalData):
    return globalData[fieldGlobalColumnsNames['porisity']] * globalData[fieldGlobalColumnsNames['ppv']] * globalData[fieldGlobalColumnsNames['pre_irrigation_ppv']] * data[row][12] * 1000


@tryDecorator
def Pfunc(data, row, globalData):
    return data[row][13] * globalData[fieldGlobalColumnsNames['max_ppv']] - data[row][14]


@tryDecorator
def Qfunc(data, row, globalData):
    return data[row][13] * globalData[fieldGlobalColumnsNames['start_ppv']]


@tryDecorator
def Rfunc(data, row):
    if not row: return 0
    return data[row][16] - data[row-1][16]


@tryDecorator
def Sfunc(data, row):
    if not row:
        return data[row][16]
    return data[row-1][25]


@tryDecorator
def Tfunc(data, row):
    return data[row][13] - data[row][18] + data[row][11]

@tryDecorator
def Wfunc(data, row):
    if data[row][20] + data[row][21] < data[row][19]:
        return data[row][20] + data[row][21]
    if data[row][20] + data[row][21] > data[row][19]:
        return data[row][19]
    return data[row][20] + data[row][21]


@tryDecorator
def Xfunc(data, row):
    return data[row][18] + data[row][22] + data[row][17] - data[row][11]


@tryDecorator
def Yfunc(data, row):
    if data[row][23] > data[row][14]:
        return 0
    if data[row][23] < data[row][14]:
        return data[row][15] / 2


@tryDecorator
def Zfunc(data, row):
    return data[row][23] + data[row][24]


@tryDecorator
def AAfunc(data, row):
    return data[row][13] + data[row][17] - data[row][25]


@tryDecorator
def ABfunc(data, row):
    return -(data[row][13] + data[row][17] - data[row][14])


@tryDecorator
def ACfunc(data, row):
    return -(data[row][13] + data[row][17] - data[row][25])


def calcAllData(data, globalData, timestamp=False):
    for row in range(len(data)):
        data[row][2] = Cfunc(data, row, globalData)
        data[row][6] = round(Gfunc(data, row), 2)
        data[row][7] = round(Hfunc(data, row), 2)
        data[row][8] = round(Ifunc(data, row), 2)
        data[row][9] = round(Jfunc(data, row), 2)
        data[row][11] = round(Lfunc(data, row), 2)
        data[row][12] = round(Mfunc(data, row), 2)
        data[row][13] = round(Nfunc(data, row, globalData), 2)
        data[row][14] = round(Ofunc(data, row, globalData), 2)
        data[row][15] = round(Pfunc(data, row, globalData), 2)
        data[row][16] = round(Qfunc(data, row, globalData), 2)
        data[row][17] = round(Rfunc(data, row), 2)
        data[row][18] = round(Sfunc(data, row), 2)
        data[row][19] = round(Tfunc(data, row), 2)
        data[row][22] = round(Wfunc(data, row), 2)
        data[row][23] = round(Xfunc(data, row), 2)
        data[row][24] = round(val, 2) if (val := Yfunc(data, row)) else 0
        data[row][25] = round(Zfunc(data, row), 2)
        data[row][26] = round(AAfunc(data, row), 2)
        data[row][27] = round(ABfunc(data, row), 2)
        data[row][28] = round(ACfunc(data, row), 2)
        data[row][1] = data[row][1] if timestamp else dateToStr(datetime.datetime.fromtimestamp(data[row][1]))
    return data

def dataFromDBtoTableData(rawData, timestamp=False):
    data = []
    columnsInRow = [0 for _ in range(29)]
    for row in range(len(rawData)):
        data.append(columnsInRow[:])
        date = datetime.datetime.fromtimestamp(int(d) if (d := rawData[row][1]) else 0)
        data[row][0] = round(d if (d := rawData[row][0]) else 0, 2)
        data[row][1] = int(d) if (d := rawData[row][1]) else 0
        # data[row][2] = round(d if (d := rawData[row][2]) else 0, 2)
        data[row][3] = round(d if (d := rawData[row][2]) else 0, 2)
        data[row][4] = round(d if (d := rawData[row][3]) else 0, 2)
        data[row][5] = round(d if (d := rawData[row][4]) else 0, 2)
        data[row][10] = round(d if (d := rawData[row][5]) else 0, 2)
        data[row][20] = round(d if (d := rawData[row][6]) else 0, 2)
        data[row][21] = round(d if (d := rawData[row][7]) else 0, 2) if type(rawData[row][7]) == float else rawData[row][7]
    return data


dashboardDates = []
lock.acquire(True)
dashboardDates = cur.execute('SELECT date from ooo_kyrsk_agroaktiv__gorshechenskiy__field62_05')
dashboardDates = dashboardDates.fetchall()
lock.release()
dashboardFields = {}
def getDashboardTable(date, request):
    
    resp = {
        'date': '',
        'header': [
            [],
            ['','','','','','','','',''],
            ['№', 'Номер поля', 'Дата сева', 'Дней с даты сева', '% FC', 'мм', 'Дней полива', 'Полив, мм', 'Осадки, мм']
        ],
        'tables': [
            []
        ] 
    }
    economy = request.json['userEconomy']
    resp['date'] = dateToStr(date)
    date = date.timestamp()
    lastMonth = -1
    for d in dashboardDates:
        resp['header'][2].append(dateToStr(datetime.datetime.fromtimestamp(d[0]))[:5])
        if lastMonth != (month := datetime.datetime.fromtimestamp(d[0]).month):
            resp['header'][1].append(MONTHS[month-1])
            lastMonth = month
            continue
        resp['header'][1].append('')

    for i in range(len(resp['header'][2])):
        resp['header'][0].append(get_column_letter(i+1))

    lock.acquire(True)
    fields = cur.execute(f'SELECT * from {economy}').fetchall() if (economy != dashboardFields.get('economy')) else dashboardFields['fields']
    lock.release()
    dashboardFields['fields'] = fields
    dashboardFields['economy'] = economy
    rowTemplate = ['' for _ in range(len(resp['header'][2]))]
    for field in fields:
        row = int(field[0])
        lock.acquire(True)
        data = calcAllData(dataFromDBtoTableData(cur.execute(f'SELECT * FROM {field[4]}').fetchall()), field, timestamp=True)
        lock.release()

        resp['tables'][0].append(rowTemplate[:])
        resp['tables'][0][row][0] = row
        resp['tables'][0][row][1] = field[fieldGlobalColumnsNames['field_name']]
        resp['tables'][0][row][2] = dateToStr(datetime.datetime.fromtimestamp(int(field[fieldGlobalColumnsNames['sowing_date']])))
        resp['tables'][0][row][3] = (date - field[fieldGlobalColumnsNames['sowing_date']]) // 86400
        resp['tables'][0][row][4] = min([round(d[25] / d[13] * 100) if d[13] != 0 else 0 for d in data][max(min(int(date//86400 - data[0][1]//86400), len(data)-1), 0)], 100)
        resp['tables'][0][row][5] = [round(d[25]) for d in data][max(min(int(date//86400 - data[0][1]//86400), len(data)-1), 0)]
        resp['tables'][0][row][6] = sum([1 for d in data if (d[21] if isinstance(d[21], (int, float)) else 0) > 0 and d[1]//86400 <= date//86400])
        resp['tables'][0][row][7] = round(sum([(d[21] if isinstance(d[21], (int, float)) else 0) for d in data if d[1]//86400 <= date//86400]), 2)
        resp['tables'][0][row][8] = round(sum([(d[20] if isinstance(d[20], (int, float)) else 0) for d in data if d[1]//86400 <= date//86400]), 2)
        for col, val in enumerate([round((d[21] if isinstance(d[21], (int, float)) else 0), 2) for d in data]):
            try:
                resp['tables'][0][row][col+9] = val
            except Exception as e:
                print(123123, e)
    
    [resp['tables'][0].insert(0, row) for row in resp['header'][::-1]]
    resp['tables'] = [[[{'value': val} for val in row] for row in resp['tables'][0]]]
    return resp


def getDataFromDB(db, field='62-05', economy='ooo_kyrsk_agroaktiv'):
    lock.acquire(True)
    cur.execute(f"SELECT field_id FROM {economy} WHERE field_name = ?", [field])
    cur.execute(f"SELECT * FROM {cur.fetchone()[0]}")
    resp = cur.fetchall()
    lock.release()
    return resp


def getGlobalDataFromDB(db, field='62-05', economy='ooo_kyrsk_agroaktiv'):
    lock.acquire(True)
    cur.execute(f"SELECT * FROM {economy} WHERE field_name = ?", [field])
    resp = cur.fetchall()
    lock.release()
    return resp


def getFieldsFromDB(db):
    lock.acquire(True)
    cur.execute(f"SELECT * FROM fields")
    resp = cur.fetchall()
    lock.release()
    return resp


def getCulturesFromDB(db):
    lock.acquire(True)
    cur.execute(f"SELECT * FROM cultures")
    resp = cur.fetchall()
    lock.release()
    return resp


def getMeteostationsFromDB(db):
    lock.acquire(True)
    cur.execute(f"SELECT * FROM meteostations")
    resp = cur.fetchall()
    lock.release()
    return resp


def getUserDataFromDB(db, login):
    lock.acquire(True)
    resp = db.cursor().execute('SELECT * FROM users WHERE login = ?', [login]).fetchall()
    lock.release()
    return resp if resp else None


def getTableData(db, field='62-05', economy='ooo_kyrsk_agroaktiv', timestamp=False):
    return calcAllData(dataFromDBtoTableData(getDataFromDB(db, field, economy), timestamp=timestamp), getGlobalDataFromDB(db, field, economy), timestamp=timestamp)


def dateToStr(date):
    try:
        resp = f'{date.day if date.day > 9 else f"0{date.day}"}.{date.month if date.month > 9 else f"0{date.month}"}.{date.year}'
    except Exception as e:
        print(e)
        resp = ''
    return resp


def strToDate(strDate):
    return datetime.datetime.fromisoformat('-'.join(reversed(strDate.split('.'))) + 'T00:00:00.000000')


def tryFloatValue(value):
    try:
        if value == '':
            return 0
        return float(value)
    except Exception as e:
        print(e)
        return None


def authorize(request):
    userData = usersDb.cursor().execute('SELECT * FROM users WHERE login_token = ?', [request.json['loginToken']]).fetchall()[0]
    print('authorization')
    print(f'\t{userData}')
    if not userData:
        print('\tauthorization failed!')
        print(f'\t{request.json}')
        return make_response('', 401)
    print('\tauthorization success!')
    return userData


@app.route('/api/login/', methods=['POST'])
@cross_origin(supports_credentials=True)
def login():
    resp = {
        'isLogined': False,
        'roles': [0],
        'loginToken': '',
        'userEconomy': '',
        'userEconomyName': '',
        'writableColumns': [],
    }
    login = request.json['data']['login']
    password = request.json['data']['password']
    userData = getUserDataFromDB(usersDb, login)
    print(login, password)
    print(userData)
    if userData[0][1] == hashlib.sha256(password.encode()).hexdigest():  
        resp['isLogined'] = True
        resp['roles'] = userData[0][5].split(',') if userData[0][5] else [0]
        resp['loginToken'] = hashlib.sha256(f'{login} {int(datetime.datetime.now().timestamp())}'.encode()).hexdigest()
        resp['userEconomy'] = userData[0][6]
        resp['userEconomyName'] = [d[0] for d in cur.execute('SELECT economy_name, economy_id from economies').fetchall() if d[1] == userData[0][6]][0]
        resp['writableColumns'] = [int(l) for l in userData[0][7].split(',')] if userData[0][7] else [-1]
        usersDb.cursor().execute('UPDATE users set login_token = ? WHERE login = ?', (resp['loginToken'], login))
        usersDb.commit()
    return make_response(jsonify(resp))


def adminpanelAddEconomy(request, userData):
    print(request.json['data']['action']['economyName'], request.json['data']['action']['economyId'])
    tableParams = 'id INTEGER PRIMARY KEY AUTOINCREMENT,' + ','.join([f'{l[1]} {l[2]}' for l in cur.execute('pragma table_info(ooo_kyrsk_agroaktiv)').fetchall()[1:]])
    cur.execute(f"INSERT INTO economies(economy_name, economy_id) values(?, ?)", (request.json['data']['action']['economyName'], request.json['data']['action']['economyId']))
    cur.execute(f"CREATE TABLE {request.json['data']['action']['economyId']}({tableParams})")
    db.commit()
    return make_response('')

def setUserColumns(request, userData):
    print(request.json['data'])
    try:
        [int(l.strip(' ')) for l in request.json['data']['action']['userColumns'].strip(', ').split(',')]
    except Exception as e:
        print('setUserColumns error: ', e)
        
    writableColumns = ','.join(list({'10', '20', '21'} & set([str(int(l.strip(', '))) for l in request.json['data']['action']['userColumns'].strip(', ').split(',') if l])))
        
    usersDb.cursor().execute("UPDATE users SET writable_columns = ? WHERE login = ?", [writableColumns, request.json['data']['action']['user']])
    usersDb.commit()
    return make_response('')

adminpanelActions = {
    'addEconomy': adminpanelAddEconomy,
    'setUserColumns': setUserColumns,
}

@app.route('/api/sendadminpanelaction/', methods=['POST'])
@cross_origin(supports_credentials=True)
def adminpanelAction():
    userData = authorize(request)
    return adminpanelActions[request.json['data']['action']['actionName']](request, userData)


def dashboardChanges(tableName, request):
    print(11121112111, request.json)
    try:
        [float(a) for a in request.json['date'].split('.')]
        if len(request.json['date']) != 10:
            print(f'date format invalid: {request.json["date"]}')
            return make_response('')
    except Exception as e:
        print(1, e)
        return make_response('')
    
    if 'changes' not in request.json['data']:
        print(strToDate(request.json['date']))
        return make_response(jsonify(getDashboardTable(strToDate(request.json['date']), request)))
    
    changes = request.json['data']['changes']
    for change in changes.values():
        economy = request.json['userEconomy']
        val, date = (change['value'], change['date'])
        lock.acquire(True)
        field = [d[1] for d in cur.execute(f'SELECT field_name, field_id from {economy}') if d[0] == change['field']][0]
        lock.release()
        date = int(strToDate(dateStr := f'{date}.{datetime.datetime.now().year}').timestamp())
        try:
            [float(a) for a in val.split('.')]
            [float(a) for a in request.json['date'].split('.')]
            1 / 10 - len(request.json['date'])
        except Exception as e:
            print(e)
            continue
        lock.acquire(True)
        print(date)
        # if not cur.execute(f'SELECT * FROM {field} WHERE date = ?', [date]).fetchall():
        #     newId = (datetime.datetime.fromtimestamp(date) - strToDate(dateStr := f'01.05.{datetime.datetime.now().year}')).days
        #     cur.execute(f'INSERT INTO {field}(id, date) VALUES(?, ?)', [newId, date])
        #     print(f'INSERT {newId} {date}')
        #     db.commit()
        print(f'writing: {val} to {dateStr}')
        print(f'UPDATE {field} set watering = ? WHERE date = ?', (float(val), date))
        cur.execute(f'UPDATE {field} set watering = ? WHERE date = ?', (float(val), date))
        db.commit()
        lock.release()
    return make_response(jsonify(getDashboardTable(strToDate(request.json['date']), request)))


def fieldChanges(tableName, request):
    tableName = tableName.split('|')[1]
    print(f'SELECT field_id FROM {request.json["userEconomy"]} WHERE field_name = "{tableName}"')
    fieldId = cur.execute(f'SELECT field_id FROM {request.json["userEconomy"]} WHERE field_name = "{tableName}"').fetchone()[0]
    for key in (changes := request.json['data']['changes']):
        row, col = [int(l.strip('rc')) for l in key.split('|')]
        value = tryFloatValue(changes[key])
        print(tableName, row, col, value)
        if value == None: 
            continue
        lock.acquire(True)
        cur.execute(f'UPDATE {fieldId} SET {fieldColumns[writableColumnsMatch[col]]} = ? WHERE id = ?', (value, row))
        db.commit()
        print(f'{value} writed into {fieldId} {fieldColumns[writableColumnsMatch[col]]} id = {row}')
        lock.release()
    
    lock.acquire(True)
    fields = cur.execute(f'SELECT * from {request.json["userEconomy"]}').fetchall()
    lock.release()
    
    field = [l for l in fields if tableName in l[3]][0]
    
    lock.acquire(True)
    data = calcAllData(dataFromDBtoTableData(cur.execute(f'SELECT * FROM {fieldId}').fetchall()), field, timestamp=False)
    lock.release()
    
    resp = [[{'value': cell if cell else ''} for cell in row] for row in data]
    resp.insert(0, tableheader)
    # pprint(resp)
    return make_response(jsonify(resp))


tableCase = {
    'dashboard': dashboardChanges,
    'field': fieldChanges
}

@app.route('/api/sendtablechanges/', methods=['POST'])
@cross_origin(supports_credentials=True)
def sendTableChanges():
    authorize(request)
    tableName = request.json['data']['tableName']
    return tableCase[tableName.split('|')[0]](tableName, request)


@app.route('/api/mapicon/')
def mapicon():
    return send_from_directory(directory='./', path='./mapicon.png')


@app.route('/api/getmapfields/', methods=['POST'])
def getMapFields():
    userData = authorize(request)
    resp = {
        'settings': {
            'zoom': 10,
            'center': {
                'lat': 51.46471, 
                'lng': 37.23724
            }
        },
        'data': []
    }
    fields = cur.execute(f'SELECT * from {userData[6]}').fetchall()
    
    date = 1663693220
    i = 0
    for _, f in enumerate(os.listdir(f'./kml/{userData[6]}/')):
        # field = f"{int(f.split('.')[0].split('-')[-2])}-{f.split('.')[0].split('-')[-1]}"
        field = [l for l in fields if '-'.join([str(int(n)) for n in l[fieldGlobalColumnsNames['field_name']].split('-')]) in '-'.join([str(int(n)) for n in f.split('.')[0].split('-')[-2:]])]
        if not field: 
            print(f)
            continue
        field = field[0]
        isAvalible = bool(field)
        data = calcAllData(dataFromDBtoTableData(cur.execute(f'SELECT * FROM {field[4]}').fetchall()), field, timestamp=True) if isAvalible else []
        resp['data'].append({
            'field': field[fieldGlobalColumnsNames['field_name']],
            'coordinates': [],
            'center': {
                'lat': 51.46471, 
                'lng': 37.23724
            },
            'isAvalible': isAvalible,
            'color': '#' + redYellowGreenFade[min([round(d[25] / d[13] * 512) if d[13] != 0 else 0 for d in data][max(min(int(date//86400 - data[0][1]//86400), len(data)-1), 0)], 511)] if isAvalible else '#000000'
        })


        with open(f'./kml/{userData[6]}/' + f, 'r', encoding='utf-8') as file:
            document = xml.parse(file)
            rawCoords = document.getElementsByTagName('coordinates')[0].childNodes[0].wholeText.strip('\n\t ').split(' ')
            coords = [{'lat': float(l.split(',')[1]), 'lng': float(l.split(',')[0])} for l in rawCoords]
            resp['data'][i]['coordinates'] = coords[:]
            resp['data'][i]['center']['lat'] = sum([lat['lat'] for lat in coords]) / len(coords)
            resp['data'][i]['center']['lng'] = sum([lng['lng'] for lng in coords]) / len(coords)
        i += 1
    return make_response(jsonify(resp))


@app.route('/api/getgraphics/', methods=['POST'])
def getGraphics():
    userData = authorize(request)
    field = request.json['data']['field']

    resp = []

    fields = cur.execute(f'SELECT * from {userData[6]}').fetchall()
    print(field)
    field = [l for l in fields if field in l[3]][0]
    data = calcAllData(dataFromDBtoTableData(cur.execute(f'SELECT * FROM {field[4]}').fetchall()), field, timestamp=True)

    for row in data:
        resp.append({
                    'data': f'{(d := datetime.datetime.fromtimestamp(row[1])).day} {MONTHS[d.month-1]}',
                    'humidityRange': round(row[27]) if isinstance(row[27], (int, float)) else None,
                    'humidity': round(row[28]) if isinstance(row[28], (int, float)) else None,
                    'waterIntake': round(row[11]) if isinstance(row[11], (int, float)) else None,
                    'rain': round(row[20]) if isinstance(row[20], (int, float)) else None,
                    'watering':  round(row[21]) if isinstance(row[21], (int, float)) else None
                })
    return make_response(jsonify(resp))


@app.route('/api/getglobaldata/', methods=['POST'])
@cross_origin(supports_credentials=True)
def getGlobalData():
    print('//////////////')
    userData = authorize(request)
    print('//////////////')
    resp = {
        'table': [],
        'selection':
        {
            'cultures': [],
            'meteostations': []
        }
    }
    
    field = request.json['data']['field']
    
    data = getGlobalDataFromDB(db, field, userData[6])[0]
    resp['selection']['cultures'] = {
        'value': {'value': data[fieldGlobalColumnsNames['culture']], 'label': data[fieldGlobalColumnsNames['culture']]},
        'list': [{'value': c[1], 'label': c[1]} for c in getCulturesFromDB(db)]
        }
    resp['selection']['meteostations'] = {
        'value': {'value': data[fieldGlobalColumnsNames['meteostation']], 'label': data[fieldGlobalColumnsNames['meteostation']]},
        'list': [{'value': m[1], 'label': m[1]} for m in getMeteostationsFromDB(db)]
        }
    resp['table'].append(tuple([dateToStr(datetime.datetime.fromtimestamp(data[10]))]) + data[11:] + data[8:10] + tuple([data[7]]))
    print('||||||||||||')
    pprint(resp)
    return make_response(jsonify(resp))

columnsMatch = {
    0: 10,
    1: 11,
    2: 12,
    3: 13,
    4: 14,
    5: 15,
    6: 8,
    7: 9,
    8: 7,
}
@app.route('/api/sendglobaldatachange/', methods=['POST'])
@cross_origin(supports_credentials=True)
def sendGlobalDataChange():
    userData = authorize(request)
    field = request.json['data']['field']

    for change in request.json['data']['changes']:
        value = request.json['data']['changes'][change]
        col = int(change.split(',')[0]) + 1
        lock.acquire(True)
        print(col, value)
        isWriting = True
        try:
            if col not in [7, 8]:
                [float(a) for a in value.split('.')]
                strToDate(value) if col == 1 else ''
        except Exception as e:
            print(e)
            isWriting = False

        if (col != 1 or len(value) == 10) and isWriting:
            print(f'UPDATE {userData[6]} set {fieldGlobalColumns[columnsMatch[col-1]]} = ? WHERE field_name = {field}', [value if col != 1 else int(strToDate(value).timestamp())])
            cur.execute(f'UPDATE {userData[6]} set {fieldGlobalColumns[columnsMatch[col-1]]} = ? WHERE field_name = "{field}"', [value if col != 1 else int(strToDate(value).timestamp())])
            db.commit()
            print('writed!')
        lock.release()
    return make_response('')


@app.route('/api/senddashboardtablechanges/', methods=['POST'])
@cross_origin(supports_credentials=True)
def sendDashboardTableChanges():
    try:
        [float(a) for a in request.json['date'].split('.')]
        if len(request.json['date']) != 10:
            print(f'date format invalid: {request.json["date"]}')
            return make_response('')
    except Exception as e:
        print(e)
        return make_response('')
    if 'change' not in request.json:
        print(strToDate(request.json['date']))
        return make_response(jsonify(getDashboardTable(strToDate(request.json['date']))))
    change = request.json['change']
    val, date, field = (change['val'], change['date'], change['field'])
    date = int(strToDate(dateStr := f'{date}.{datetime.datetime.now().year}').timestamp())
    try:
        [float(a) for a in val.split('.')]
        [float(a) for a in request.json['date'].split('.')]
        1 / 10 - len(request.json['date'])
    except Exception as e:
        print(e)
        return make_response('')
    lock.acquire(True)
    print(f'writing: {val} to {dateStr}')
    cur.execute(f'UPDATE field{field.replace("-", "z")} set watering_count = ? WHERE date = ?', (float(val), date))
    db.commit()
    lock.release()
    return make_response(jsonify(getDashboardTable(strToDate(request.json['date']))))


tableheader = [{'value': '№ п/п'}, {'value': 'Дата'}, {'value': 'день от даты сева'}, {'value': 'температура воздуха t, °С'}, {'value': 'влажность воздуха А, %'}, {'value': 'скорость ветра V2 на высоте 2м, м/с'}, {'value': 'испаряемость Ei, мм/сут'}, {'value': 'средне многолетняя суточная испаряемость Ео, мм/сут'}, {'value': 'Сумма температур воздуха с даты сева ∑t, °С'}, {'value': 'средне многолетний Кбо'}, {'value': 'текущий Кбi'}, {'value': 'Водопотреб ление Ev, мм/сут '}, {'value': 'Расчетный слой почвы h, м'}, {'value': 'Влагоемкость расчетного слоя почвы слоя почвы Wнв, мм'}, {'value': 'Предполивные влагозапасы расчетного слоя слоя почвы Wпр, мм'}, {'value': 'Макс. возможная поливная норма m, мм'}, {'value': 'Весенние влагозапасы на 01.05 Wв=0,9Wнв,мм '}, {'value': 'Приращение Wв'}, {'value': 'Начальные влагозапасы Wн, мм'}, {'value': '(Wнв-Wн)+Ev'}, {'value': 'Атмосферные осадки Р, мм'}, {'value': 'Реализованный полив mф , мм'}, {'value': 'Эффективное поступление влаги Нэф, мм'}, {'value': 'Конечные влагозапасы в расчетном слое почвы Wк,мм '}, {'value': 'Расчетный полив mр, мм'}, {'value': 'Конечные влагозапасы в расчетном слое почвы после полива, мм'}, {'value': 'сколько мм влаги не хватает до верхнего порога (0,95 ППВ)?'}, {'value': 'Диапазон, мм'}, {'value': 'Влагозапасы факт, мм'}]
@app.route('/api/gettable/', methods=['POST'])
@cross_origin(supports_credentials=True)
def getTable():
    userData = authorize(request)
    field = request.json['data']['field']

    resp = [[{'value': cell if cell else ''} for cell in row] for row in getTableData(db, field, userData[6])]
    resp.insert(0, tableheader)
    # pprint(resp)
    return make_response(jsonify(resp))


@app.route('/api/getdashboardtable/', methods=['POST'])
@cross_origin(supports_credentials=True)
def getDashboardTableData():
    return make_response(jsonify(getDashboardTable(datetime.datetime.now(), request)))


# @app.route('/api/gettemplate/')
# def getTemplate():
    field = request.args['field']
    wb = ox.load_workbook('template.xlsx')
    ws = wb['template']
    data = getDataFromDB(getConnection(), field)
    for row, rowVal in enumerate(data):
        for col, colVal in enumerate(rowVal):
            if col == 1:
                date = datetime.datetime.fromtimestamp(int(colVal))
                colVal = '.'.join([str(d) if (d := date.day) > 9 else f'0{d}', str(m) if (m := date.month) > 9 else f'0{m}', str(date.year)])
            ws[row+2][col].value = colVal
    wb.save(f'fieldsTemplates/field{field}.xlsx')
    wb.close()
    return send_file(f'fieldsTemplates/field{field}.xlsx')


# @app.route('/api/sendtemplate/', methods=['POST'])
# def sendTemplate(field='62-05'):
    timestamp = int(datetime.datetime.now().timestamp())
    filename = f'./fieldsRecivedTemlates/{timestamp}.xlsx'
    request.files.to_dict()['files[]'].save(filename)
    wb = ox.load_workbook(filename)
    ws = wb.worksheets[0]
    data = [
        (r[0].value, 
        int(strToDate(r[1].value).timestamp()),
        r[2].value, 
        r[3].value, 
        r[4].value, 
        r[5].value, 
        r[6].value, 
        r[7].value, 
        r[8].value
        ) for r in [row for row in ws.rows][1:] if r[1].value]
    for rowId, rowVal in [(d[0], d[1:]) for d in data]:
        for col, colValue in enumerate(rowVal):
            lock.acquire(True)
            cur.execute(f'UPDATE field{field.replace("-", "z")} set {fieldColumns[col+1]} = ? WHERE id = ?', (colValue, rowId))
            lock.release()
    db.commit()
    
    return jsonify(getTableData(db, field))


if __name__ == "__main__":
    app.run(host='0.0.0.0')


# data = [(r[0].value, int(datetime.datetime.timestamp(r[1].value)), r[2].value, r[3].value, r[4].value, r[5].value, r[10].value, r[20].value, r[21].value) for r in [row for row in ws.rows][5:] if r[1].value]

@app.route('/api/settablechange/')
def setTableChange():
    isWriting = True
    try:
        field = request.args['field']
        col = writableColumnsMatch.get(int(request.args['column']))
        val = request.args['value']
        row = request.args['row']
        [float(a) for a in val.split('.')]
        strToDate(val) if col == 2 else ''
    except Exception as e:
        print(e)
        isWriting = False
    if col and col <= 8 and val and isWriting:
        print(f"replacing {col} {row} with {val}")
        value = float(val.replace(',', '.').split('\n')[0])
        lock.acquire(True)
        cur.execute(f'UPDATE field{field.replace("-", "z")} set {fieldColumns[col-1]} = ? WHERE id = ?', (value if int(col) != 1 else int(strToDate(value).timestamp()), int(row)+1))
        lock.release()
        db.commit()
        return make_response(jsonify(getTableData(db, field)))
    return make_response('')


@app.route('/api/sendsettigstablechanges/', methods=['POST'])
def sendSettingsTableChanges(field='62-05'):
    print('settigstablechanges')
    print(request.json)
    for key in request.json:
        table, column, row = key.split(',')
        lock.acquire(True)
        if int(table):
            print(column, row, request.json[key])
            cur.execute(f'UPDATE field{field.replace("-", "z")} set {fieldColumns[int(column)+1]} = ? WHERE id = ?', (request.json[key], int(row)+1))
            continue
        cur.execute(f'UPDATE field{field.replace("-", "z")}global set {fieldGlobalColumns[int(row)+1]} = ? WHERE id = 0', [request.json[key] if int(row) != 0 else int(strToDate(request.json[key]).timestamp())])
        lock.release()
    db.commit()
    return make_response('')


@app.route('/api/getsettingstable/')
@cross_origin(supports_credentials=True)
def getSettingsTable(field='62-05'):
    data = []
    data.append([[d] for d in getGlobalDataFromDB(db, field)])
    data[0][0][0] = dateToStr(datetime.datetime.fromtimestamp(data[0][0][0]))
    data.append([
        (
            dateToStr(datetime.datetime.fromtimestamp(d[fieldColumnsNames['date']])), 
            round(d[fieldColumnsNames['air_temp']], 2),
            round(d[fieldColumnsNames['air_hum']], 2), 
            round(d[fieldColumnsNames['wind_speed']], 2)
        ) for d in getDataFromDB(db, field)])
    return make_response(jsonify(data))
